"""Presigned URL upload flow (BRD v8) — 3-step: request URL → upload S3 → confirm."""
import io
import logging
from datetime import datetime, timezone

from fastapi import HTTPException, status
from sqlalchemy.orm import Session

from app.core import storage
from app.core.config import settings
from app.core.email_service import send_request_confirmation
from app.core.report_client import check_health as check_report_service_health
from app.models.etariff_usage_log import ETariffUsageLog
from app.models.request import Request, RequestFile, RequestStatus, RequestType
from app.models.user import User, UserRole
from app.requests import repository
from app.requests.tasks import run_ai_analysis

logger = logging.getLogger(__name__)

ALLOWED_EXTENSIONS = {".xlsx", ".xlsb"}
MAX_FILE_SIZE = settings.MAX_UPLOAD_SIZE_MB * 1024 * 1024


def _check_ai_service():
    if not check_report_service_health():
        raise HTTPException(status.HTTP_503_SERVICE_UNAVAILABLE, "Hệ thống AI đang bảo trì, vui lòng thử lại sau.")


def _validate_extension(filename: str):
    ext = f".{filename.rsplit('.', 1)[-1].lower()}" if "." in filename else ""
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"File '{filename}' not allowed. Only .xlsx/.xlsb accepted.")


def _build_display_id(db: Session, tenant_code: str, tenant_id: int) -> str:
    count = repository.count_by_tenant(db, tenant_id)
    return f"{tenant_code}-{str(count + 1).zfill(3)}"


def request_presigned_url(
    db: Session, filename: str, file_size: int, request_type: str,
    chc_modules: list[str] | None, user: User,
) -> dict:
    """Step 1: Create request record + return presigned upload URL."""
    _check_ai_service()
    _validate_extension(filename)

    if file_size > MAX_FILE_SIZE:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"File exceeds {settings.MAX_UPLOAD_SIZE_MB}MB limit.")

    req_type = RequestType(request_type)
    if req_type == RequestType.etariff_batch:
        _check_etariff_limit(db, user)

    display_id = _build_display_id(db, user.tenant.tenant_code, user.tenant_id)
    req = repository.create_request(
        db, user.tenant_id, user.id, display_id,
        type=req_type,
        chc_modules=chc_modules if req_type == RequestType.chc else None,
    )

    s3_key = f"{user.tenant_id}/requests/{req.id}/{filename}"
    rf = repository.create_file(db, req.id, filename, s3_key, file_size)

    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    upload_url = storage.generate_presigned_upload_url(s3_key, content_type)

    return {
        "request_id": req.id,
        "file_id": rf.id,
        "display_id": req.display_id,
        "upload_url": upload_url,
        "s3_key": s3_key,
        "expires_in": 900,
    }


def confirm_upload(db: Session, request_id: int, user: User) -> dict:
    """Step 3: Validate uploaded file, count rows (batch), start AI processing."""
    req = repository.get_by_id(db, request_id)
    if not req or req.user_id != user.id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Request not found.")
    if req.status != RequestStatus.pending:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Request already confirmed.")

    rf = req.files[0] if req.files else None
    if not rf or not rf.s3_key:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "No file found for this request.")

    row_count = None
    quota_remaining = None

    # Batch mode: count rows and check quota
    if req.type == RequestType.etariff_batch:
        row_count = _count_excel_rows(rf.s3_key)
        quota_remaining = _get_quota_remaining(db, user)

        if quota_remaining <= 0:
            raise HTTPException(status.HTTP_429_TOO_MANY_REQUESTS, "Hết lượt tra cứu hôm nay.")

        # Log usage
        _log_etariff_usage(db, user, req, "batch", row_count, rf.original_filename)

    # Transition: pending → ai_processing
    req.status = RequestStatus.ai_processing
    req.ai_processing_started_at = datetime.now(timezone.utc)
    db.commit()

    # Trigger AI
    run_ai_analysis.delay(rf.id)

    # Emails
    send_request_confirmation(user, req)
    _notify_admins(db, user.tenant_id, req)

    result = {
        "request_id": req.id,
        "display_id": req.display_id,
        "status": req.status.value,
    }
    if row_count is not None:
        result["row_count"] = row_count
        result["quota_remaining"] = max(0, quota_remaining - row_count)
        if row_count > quota_remaining:
            result["warning"] = f"File có {row_count} dòng nhưng bạn chỉ còn {quota_remaining} lượt."

    return result


def _count_excel_rows(s3_key: str) -> int:
    """Download file from S3 and count data rows (exclude header)."""
    try:
        from openpyxl import load_workbook
        file_bytes = storage.download_file_bytes(s3_key)
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        row_count = max(0, ws.max_row - 1)  # exclude header
        wb.close()
        return row_count
    except Exception as e:
        logger.warning("Failed to count rows for %s: %s", s3_key, e)
        return 0


def _get_quota_remaining(db: Session, user: User) -> int:
    from sqlalchemy import func
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    used = db.query(func.coalesce(func.sum(ETariffUsageLog.row_count), 0)).filter(
        ETariffUsageLog.tenant_id == user.tenant_id,
        ETariffUsageLog.created_at >= today_start,
    ).scalar() or 0
    limit = user.tenant.etariff_daily_limit if user.tenant else 10
    return limit - used


def _log_etariff_usage(db: Session, user: User, req: Request, mode: str, row_count: int, summary: str | None):
    log = ETariffUsageLog(
        user_id=user.id, tenant_id=user.tenant_id, request_id=req.id,
        mode=mode, row_count=row_count, query_summary=summary,
    )
    db.add(log)
    db.commit()


def _check_etariff_limit(db: Session, user: User):
    remaining = _get_quota_remaining(db, user)
    if remaining <= 0:
        limit = user.tenant.etariff_daily_limit if user.tenant else 10
        raise HTTPException(status.HTTP_429_TOO_MANY_REQUESTS, f"Hết lượt tra cứu E-Tariff hôm nay ({limit}/ngày).")


def _notify_admins(db: Session, tenant_id: int, req: Request):
    from app.core.email_service import send_admin_new_request
    admins = db.query(User).filter(User.tenant_id == tenant_id, User.role == UserRole.tenant_admin, User.is_active == True).all()
    for admin in admins:
        send_admin_new_request(admin, req)
